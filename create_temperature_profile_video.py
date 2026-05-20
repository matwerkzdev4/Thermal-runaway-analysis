from __future__ import annotations

import math
import shutil
import subprocess
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
from PIL import Image, ImageDraw, ImageFont

from create_clean_test_configuration_schematic import (
    BLUE,
    BLUE_EDGE,
    CELLS,
    HEATERS,
    INK,
    JACKETS,
    RED,
    TC_POINTS,
    TC_R,
)


ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Test Results.xlsm"
FRAMES_DIR = ROOT / "report_figures" / "temperature_profile_frames_v4"
OUTPUT = ROOT / "report_figures" / "thermal_runaway_temperature_profile.mp4"
FFMPEG = ROOT / "node_modules" / "ffmpeg-static" / "ffmpeg.exe"

WIDTH = 1920
HEIGHT = 1080
FPS = 15
DURATION_SECONDS = 75
END_TIME_SECONDS = 80 * 60
FRAME_COUNT = FPS * DURATION_SECONDS
TEMP_MIN = 20
TEMP_MAX = 800

EVENTS = [
    ("C2 TR", "T7", 38 * 60 + 18, (220, 55, 55)),
    ("Heaters off", None, 38 * 60 + 22, (80, 80, 80)),
    ("C1 TR", "T3", 48 * 60 + 54, (65, 125, 210)),
    ("C3 TR", "T11", 51 * 60 + 2, (230, 120, 45)),
    ("C4 TR", "T15", 63 * 60 + 27, (210, 165, 35)),
    ("C5 TR", "T19", 74 * 60 + 7, (190, 65, 185)),
]


def fmt_time(seconds: float) -> str:
    seconds = int(round(seconds))
    return f"{seconds // 60:02d}:{seconds % 60:02d}"


def load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    candidates = [
        Path("C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default(size=size)


FONT_TITLE = load_font(38, bold=True)
FONT_SUB = load_font(24)
FONT_LABEL = load_font(24, bold=True)
FONT_SMALL = load_font(18)
FONT_TINY = load_font(15)


def temp_color(temp: float | None) -> tuple[int, int, int]:
    if temp is None or math.isnan(temp):
        return (150, 150, 150)
    stops = [
        (20, (36, 95, 190)),
        (100, (55, 190, 230)),
        (200, (255, 230, 70)),
        (350, (255, 140, 35)),
        (550, (220, 45, 35)),
        (800, (245, 245, 245)),
    ]
    if temp <= stops[0][0]:
        return stops[0][1]
    for (t0, c0), (t1, c1) in zip(stops, stops[1:]):
        if temp <= t1:
            f = (temp - t0) / (t1 - t0)
            return tuple(int(c0[i] + (c1[i] - c0[i]) * f) for i in range(3))
    return stops[-1][1]


def read_data():
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True, keep_vba=True)
    ws = wb.worksheets[0]
    headers = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    idx = {header: i for i, header in enumerate(headers)}
    rows = [row for row in ws.iter_rows(min_row=3, values_only=True) if row[0] is not None]
    start = datetime.strptime(str(rows[0][0]), "%Y-%m-%d %H:%M:%S")

    times = []
    values = {tc: [] for tc in TC_POINTS if tc in idx}
    for row in rows:
        stamp = datetime.strptime(str(row[0]), "%Y-%m-%d %H:%M:%S")
        elapsed = (stamp - start).total_seconds()
        if 0 <= elapsed <= END_TIME_SECONDS:
            times.append(elapsed)
            for tc in values:
                values[tc].append(row[idx[tc]])

    times_np = np.array(times, dtype=float)
    values_np = {tc: np.array(vals, dtype=float) for tc, vals in values.items()}
    return times_np, values_np


def value_at(times: np.ndarray, values: dict[str, np.ndarray], t: float) -> dict[str, float]:
    return {tc: float(np.interp(t, times, series)) for tc, series in values.items()}


def draw_centered(draw: ImageDraw.ImageDraw, xy, text: str, font, fill):
    bbox = draw.textbbox((0, 0), text, font=font)
    draw.text((xy[0] - (bbox[2] - bbox[0]) / 2, xy[1] - (bbox[3] - bbox[1]) / 2), text, font=font, fill=fill)


def draw_layout(draw: ImageDraw.ImageDraw):
    draw.rectangle((0, 0, WIDTH, HEIGHT), fill=(247, 249, 252))
    draw.text((54, 54), "Thermal runaway test temperature profile", font=FONT_TITLE, fill=(24, 32, 42))
    draw.text((56, 98), "Animated from raw thermocouple data in Test Results.xlsm", font=FONT_SMALL, fill=(86, 98, 112))
    draw.rounded_rectangle((45, 145, 1875, 785), radius=12, fill=(238, 243, 248), outline=(215, 222, 232), width=2)

    for box in JACKETS:
        draw.rectangle(box, fill=(17, 17, 17))
    for box in HEATERS:
        draw.rectangle(box, fill=RED)

    for name, rect in CELLS.items():
        x0, y0, x1, y1 = rect
        draw.rounded_rectangle(rect, radius=6, fill=BLUE, outline=BLUE_EDGE, width=3)
        draw_centered(draw, ((x0 + x1) / 2, (y0 + y1) / 2), name, FONT_TITLE, (255, 255, 255))

    draw.text((56, 825), "Color scale", font=FONT_SMALL, fill=(24, 32, 42))
    x0, y0 = 170, 830
    for i in range(360):
        temp = TEMP_MIN + (TEMP_MAX - TEMP_MIN) * i / 359
        draw.line((x0 + i, y0, x0 + i, y0 + 24), fill=temp_color(temp))
    for label, x in [("20", x0), ("200 °C", x0 + 83), ("800", x0 + 360)]:
        draw.text((x - 12, y0 + 30), label, font=FONT_TINY, fill=(24, 32, 42))


def draw_timeline(draw: ImageDraw.ImageDraw, current_t: float):
    x0, x1, y = 590, 1810, 850
    draw.text((590, 815), "TR event timeline", font=FONT_SMALL, fill=(24, 32, 42))
    draw.line((x0, y, x1, y), fill=(188, 197, 208), width=4)
    progress = min(max(current_t / END_TIME_SECONDS, 0), 1)
    draw.line((x0, y, x0 + progress * (x1 - x0), y), fill=(230, 158, 34), width=5)
    draw.ellipse((x0 + progress * (x1 - x0) - 8, y - 8, x0 + progress * (x1 - x0) + 8, y + 8), fill=(24, 32, 42))
    for label, tc, event_t, color in EVENTS:
        ex = x0 + event_t / END_TIME_SECONDS * (x1 - x0)
        draw.line((ex, y - 28, ex, y + 28), fill=color, width=3)
        if label == "Heaters off":
            compact = "Heat off"
            label_x = ex - 48
            label_y = y + 64
        elif label == "C2 TR":
            compact = "C2 TR"
            label_x = ex - 38
            label_y = y + 34
        elif label == "C1 TR":
            compact = "C1 TR"
            label_x = ex - 38
            label_y = y + 64
        elif label == "C5 TR":
            compact = "C5 TR"
            label_x = ex - 32
            label_y = y + 64
        else:
            compact = label
            label_x = ex - 32
            label_y = y + 34
        draw.text((label_x, label_y), compact, font=FONT_TINY, fill=color)


def draw_frame(frame_no: int, times: np.ndarray, values: dict[str, np.ndarray]):
    current_t = END_TIME_SECONDS * frame_no / (FRAME_COUNT - 1)
    current = value_at(times, values, current_t)
    img = Image.new("RGB", (WIDTH, HEIGHT), (247, 249, 252))
    draw = ImageDraw.Draw(img)
    draw_layout(draw)

    draw.text((1420, 48), f"Elapsed time  {fmt_time(current_t)}", font=FONT_SUB, fill=(24, 32, 42))
    draw.text((1422, 80), "200 °C threshold shown by yellow/amber transition", font=FONT_TINY, fill=(86, 98, 112))

    for label, tc, event_t, color in EVENTS:
        if current_t >= event_t:
            idx = EVENTS.index((label, tc, event_t, color))
            draw.rounded_rectangle((56, 135 + 36 * idx, 330, 163 + 36 * idx), radius=5, fill=(255, 255, 255), outline=color, width=2)
            draw.text((68, 139 + 36 * idx), f"{label}: {fmt_time(event_t)}", font=FONT_SMALL, fill=color)

    for tc, (x, y) in TC_POINTS.items():
        temp = current.get(tc)
        color = temp_color(temp)
        radius = 16 if temp < 200 else 22 if temp < 350 else 28
        outline = (255, 255, 255) if temp >= 200 else INK
        if any(tc == event_tc and abs(current_t - event_t) < 20 for _, event_tc, event_t, _ in EVENTS if event_tc):
            radius = 34
            outline = (255, 255, 255)
        draw.ellipse((x - radius, y - radius, x + radius, y + radius), fill=color, outline=outline, width=3)
        draw_centered(draw, (x, y), tc, FONT_TINY, (0, 0, 0))
        draw.rounded_rectangle((x - 28, y + radius + 4, x + 32, y + radius + 24), radius=3, fill=(255, 255, 255), outline=(215, 222, 232), width=1)
        draw.text((x - 24, y + radius + 5), f"{temp:.0f}°C", font=FONT_TINY, fill=(24, 32, 42))

    draw_timeline(draw, current_t)
    return img


def main():
    times, values = read_data()
    if FRAMES_DIR.exists():
        shutil.rmtree(FRAMES_DIR)
    FRAMES_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    for frame_no in range(FRAME_COUNT):
        frame = draw_frame(frame_no, times, values)
        frame.save(FRAMES_DIR / f"frame_{frame_no:04d}.png", optimize=False)
        if frame_no % 100 == 0:
            print(f"rendered {frame_no}/{FRAME_COUNT}")

    if not FFMPEG.exists():
        raise FileNotFoundError(f"ffmpeg not found at {FFMPEG}")

    cmd = [
        str(FFMPEG),
        "-y",
        "-framerate",
        str(FPS),
        "-i",
        str(FRAMES_DIR / "frame_%04d.png"),
        "-c:v",
        "libx264",
        "-pix_fmt",
        "yuv420p",
        "-movflags",
        "+faststart",
        str(OUTPUT),
    ]
    subprocess.run(cmd, check=True)
    print(f"saved {OUTPUT}")


if __name__ == "__main__":
    main()
